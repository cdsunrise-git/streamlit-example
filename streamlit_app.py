import streamlit as st
from openpyxl import load_workbook, Workbook
import os

path_curr = os.path.dirname(__file__)

union_file = path_curr + "/" + "Weekly_re.xlsx"


wb = Workbook()

ws = wb.create_sheet("a")

ws['A1'].value = "test입니다."

st.error("파일명: ", union_file)

wb_dst.save(union_file)
wb_dst.close()
